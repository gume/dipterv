#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import importlib
import json
import logging
import os
import re
import statistics
import unicodedata
import urllib.error
import urllib.request
import warnings
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader


warnings.filterwarnings("ignore", message=r"Ignoring wrong pointing object.*")
logging.getLogger("pypdf").setLevel(logging.ERROR)


FORMAL_RUBRICS = [
    {
        "id": "heading_detected",
        "group": "formal",
        "description": "The document contains a recognizable assignment heading.",
        "pass_rule": "Heading includes thesis-assignment markers in Hungarian or English.",
    },
    {
        "id": "heading_spelling_ok",
        "group": "formal",
        "description": "The assignment heading matches expected wording without obvious typos.",
        "pass_rule": "Detected heading contains a canonical heading marker.",
    },
    {
        "id": "student_name_detected",
        "group": "formal",
        "description": "The student name can be extracted from the template.",
        "pass_rule": "A student marker line is found and a name-like line is extracted.",
    },
    {
        "id": "title_detected",
        "group": "formal",
        "description": "The thesis title can be extracted.",
        "pass_rule": "A non-template title line is found near the top or derived from filename.",
    },
    {
        "id": "advisor_detected",
        "group": "formal",
        "description": "The departmental supervisor is listed.",
        "pass_rule": "A supervisor/advisor label is detected in HU/EN text.",
    },
    {
        "id": "date_detected",
        "group": "formal",
        "description": "The location and date line is present.",
        "pass_rule": "A location + date expression is detected.",
    },
    {
        "id": "page_count_reasonable",
        "group": "formal",
        "description": "The PDF length matches the expected assignment format.",
        "pass_rule": "Page count is between 1 and 2.",
    },
    {
        "id": "task_list_present",
        "group": "formal",
        "description": "The assignment contains an explicit task list.",
        "pass_rule": "At least 4 bullet items are detected.",
    },
]

CONTENT_RUBRICS = [
    {
        "id": "objective_paragraph_present",
        "group": "content",
        "description": "The task description has a readable objective/problem statement.",
        "pass_rule": "Introductory text before the task list has at least 300 characters.",
    },
    {
        "id": "literature_task_present",
        "group": "content",
        "description": "The work includes literature review or background study.",
        "pass_rule": "Language-specific literature keywords are present.",
    },
    {
        "id": "implementation_task_present",
        "group": "content",
        "description": "The work includes implementation, modeling, or experimental work.",
        "pass_rule": "Language-specific implementation keywords are present.",
    },
    {
        "id": "evaluation_task_present",
        "group": "content",
        "description": "The work includes evaluation, comparison, or validation.",
        "pass_rule": "Language-specific evaluation keywords are present.",
    },
    {
        "id": "report_task_present",
        "group": "content",
        "description": "The work includes a written report or presentation deliverable.",
        "pass_rule": "Language-specific deliverable keywords are present.",
    },
    {
        "id": "task_count_reasonable",
        "group": "content",
        "description": "The task list is neither too short nor too long.",
        "pass_rule": "Bullet item count is between 4 and 8.",
    },
]

NOISE_LINE_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"Budapesti Muszaki es Gazdasagtudomanyi Egyetem",
        r"Villamosmernoki es Informatikai Kar",
        r"Tavkozlesi es Mesterseges Intelligencia Tanszek",
        r"Magyar tudosok krt",
        r"Department of Telecommunications",
        r"Faculty of Electrical Engineering",
        r"https?://",
        r"E-mail:",
        r"Tel\\.:",
        r"Web:",
    ]
]

LANGUAGE_PATTERNS = {
    "hu": (
        "hallgato",
        "feladat",
        "konzulens",
        "tanszek",
        "budapest",
        "szakdolgozat",
        "diplomatervezesi",
        "irodalom",
        "ertekel",
    ),
    "en": (
        "student",
        "thesis",
        "assignment",
        "advisor",
        "supervisor",
        "department",
        "evaluation",
        "literature",
        "report",
    ),
}

KEYWORD_GROUPS = {
    "hu": {
        "literature": ("irodalom", "irodalomkutatas", "attekintes", "szakirodalom"),
        "implementation": ("fejleszt", "keszit", "implement", "modell", "tervez", "kialakit"),
        "evaluation": ("ertekel", "osszehasonlit", "vizsgalat", "validal", "mer", "elemz"),
        "report": ("irasos beszamolo", "osszefoglalo", "prezentacio", "szobeli"),
    },
    "en": {
        "literature": ("literature", "survey", "review", "state of the art", "background"),
        "implementation": ("implement", "develop", "design", "build", "model", "prototype"),
        "evaluation": ("evaluate", "validation", "compare", "benchmark", "analysis", "test"),
        "report": ("report", "documentation", "presentation", "written summary"),
    },
}

HEADING_MARKERS = (
    "diplomatervezesi feladat",
    "diplomaterv feladatkiiras",
    "diplomaterv feladat",
    "szakdolgozati feladat",
    "szakdolgozat feladat",
    "thesis assignment",
    "bsc thesis assignment",
    "msc thesis assignment",
    "thesis task description",
    "bsc thesis task description",
    "msc thesis task description",
    "msc/bsc thesis task description",
    "project assignment",
)

STUDENT_NAME_EXCLUDE_MARKERS = (
    *HEADING_MARKERS,
    "thesis task description",
    "msc thesis task description",
    "bsc thesis task description",
    "msc/bsc thesis task description",
    "feladatkiiras",
)

STUDENT_MARKERS = (
    "hallgato reszere",
    "for student",
    "for the student",
    "student:",
    "candidate for",
)

ADVISOR_PATTERNS = [
    r"Tanszeki\s+konzulens\s*:\s*(.+)",
    r"Konzulens\s*:\s*(.+)",
    r"Supervisor\s*:\s*(.+)",
    r"Advisor\s*:\s*(.+)",
    r"Departmental\s+supervisor\s*:\s*(.+)",
]

DATE_PATTERNS = [
    r"Budapest,\s*[0-9]{4}\.\s*[A-Za-zA-Z\u00C0-\u017F]+\s+[0-9]{1,2}\.",
    r"Budapest,\s*[A-Za-zA-Z\u00C0-\u017F]+\s+[0-9]{1,2},\s*[0-9]{4}",
    r"[A-Za-zA-Z\u00C0-\u017F ]+,\s*[A-Za-zA-Z\u00C0-\u017F]+\s+[0-9]{1,2},\s*[0-9]{4}",
]


@dataclass
class SpellcheckSummary:
    status: str
    issue_count: int
    grammar_issue_count: int
    spelling_issue_count: int
    summary: str


@dataclass
class LlmCheckSummary:
    status: str
    score: int
    verdict: str
    summary: str
    parsed_json: str = ""
    raw_answer: str = ""


def strip_accents(text: str) -> str:
    mapping = str.maketrans(
        {
            "a": "a",
            "A": "A",
            "á": "a",
            "Á": "A",
            "é": "e",
            "É": "E",
            "í": "i",
            "Í": "I",
            "ó": "o",
            "Ó": "O",
            "ö": "o",
            "Ö": "O",
            "ő": "o",
            "Ő": "O",
            "ú": "u",
            "Ú": "U",
            "ü": "u",
            "Ü": "U",
            "ű": "u",
            "Ű": "U",
        }
    )
    return text.translate(mapping)


def normalize_whitespace(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]*\n[ \t]*", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_layout_spacing(text: str) -> str:
    text = normalize_whitespace(text)
    if not text:
        return text

    def merge_title_continuation_lines(source: str) -> str:
        lines = source.split("\n")
        if len(lines) < 3:
            return source

        starters = {"and", "for", "with", "in", "of", "to", "on", "using", "towards", "toward"}
        merged: list[str] = []
        index = 0
        # Restrict to early document region where title block appears.
        scan_limit = min(len(lines), 40)

        while index < len(lines):
            if index + 2 < scan_limit:
                first = lines[index].strip()
                middle = lines[index + 1].strip()
                third = lines[index + 2].strip()

                if first and not middle and third:
                    words = third.split()
                    first_word = words[0].lower() if words else ""
                    title_like = (
                        len(words) <= 5
                        and third[:1].islower()
                        and (first_word in starters or len(words) <= 3)
                        and not first.endswith((".", "!", "?", ":", ";"))
                        and not re.match(r"^[\u2022\u25CF\uF0B7\-*\u2023]", first)
                        and not re.match(r"^[\u2022\u25CF\uF0B7\-*\u2023]", third)
                    )
                    if title_like:
                        merged.append(f"{first} {third}")
                        index += 3
                        continue

            merged.append(lines[index])
            index += 1

        return "\n".join(merged)

    def merge_title_continuation_in_header(source: str) -> str:
        split_markers = [
            "Tasks to be performed by the student will include",
            "A hallgató feladatának a következőkre kell kiterjednie",
            "The task of the candidate is",
            "The following subtasks should be elaborated",
        ]
        cut_index = len(source)
        lower_source = source.lower()
        for marker in split_markers:
            idx = lower_source.find(marker.lower())
            if idx != -1:
                cut_index = min(cut_index, idx)

        header = source[:cut_index]
        body = source[cut_index:]

        continuation_pattern = re.compile(
            r"(?m)^(?P<a>[^\n]{20,180})\n\n(?P<b>(?:and|for|with|in|of|to|on|using|towards|toward)\s[^\n]{2,120})$"
        )

        for _ in range(3):
            header, replaced = continuation_pattern.subn(r"\g<a> \g<b>", header)
            if replaced == 0:
                break

        return header + body

    text = merge_title_continuation_lines(text)
    text = merge_title_continuation_in_header(text)

    # Some PDFs place multiple bullet items on one line; split them early.
    text = re.sub(r"\s+([\u2022\u25CF\uF0B7\u2023])\s+", r"\n\1 ", text)

    # Fix common merged student-role fragments (name + degree/role words).
    text = re.sub(
        (
            r"(?<=[A-Za-z\u00C0-\u017F])"
            r"(?=("
            r"[uü]zemm[ée]rn[öo]k-?informatikus|"
            r"m[ée]rn[öo]kinformatikus|"
            r"hallgat[óo]|"
            r"r[ée]sz[ée]re|"
            r"reszere"
            r")\b)"
        ),
        " ",
        text,
        flags=re.IGNORECASE,
    )

    bullet_pattern = re.compile(r"^\s*[\u2022\u25CF\uF0B7\-*\u2023]\s+")
    paragraphs = re.split(r"\n\n+", text)
    normalized_paragraphs = []

    for paragraph in paragraphs:
        lines = [line.strip() for line in paragraph.splitlines() if line.strip()]
        if not lines:
            continue

        if any(bullet_pattern.match(line) for line in lines):
            # Keep one bullet item per line and fold bullet-only marker lines.
            bullet_lines: list[str] = []
            for line in lines:
                if re.match(r"^\s*[\u2022\u25CF\uF0B7\u2023]\s*$", line):
                    bullet_lines.append("•")
                    continue
                if bullet_lines and bullet_lines[-1] == "•" and not bullet_pattern.match(line):
                    bullet_lines[-1] = f"• {line}"
                    continue
                if bullet_pattern.match(line):
                    line = re.sub(r"^\s*[\u2022\u25CF\uF0B7\-*\u2023]\s*", "• ", line)
                bullet_lines.append(line)
            normalized_paragraphs.append("\n".join(bullet_lines))
            continue

        merged = lines[0]
        for line in lines[1:]:
            prev = merged.rstrip()
            cur = line.lstrip()

            if prev.endswith("-"):
                # Hyphenation at line end: remove dash and join without extra space.
                merged = prev[:-1] + cur
                continue

            prev_last = re.findall(r"[A-Za-z\u00C0-\u017F]+$", prev)
            cur_first = re.findall(r"^[A-Za-z\u00C0-\u017F]+", cur)
            if prev_last and cur_first:
                prev_token = prev_last[0]
                cur_token = cur_first[0]
                # Join line-wrapped word fragments like "algoritm" + "usok".
                if (
                    prev_token.isalpha()
                    and cur_token.isalpha()
                    and prev_token[0].islower()
                    and prev_token[-1].islower()
                    and cur_token[0].islower()
                    and 3 <= len(prev_token) <= 8
                    and 2 <= len(cur_token) <= 8
                ):
                    merged = prev + cur
                    continue

            # Join wrapped lines inside a paragraph with a single space.
            merged = f"{prev} {cur}"

        normalized_paragraphs.append(normalize_whitespace(merged))

    # Merge title continuations split by an empty line, e.g. "... Kubernetes" + "kornyezetben".
    merged_paragraphs: list[str] = []
    for paragraph in normalized_paragraphs:
        current = normalize_whitespace(paragraph)
        if merged_paragraphs:
            previous = merged_paragraphs[-1]
            prev_words = previous.split()
            cur_words = current.split()
            if (
                1 <= len(cur_words) <= 3
                and 3 <= len(prev_words) <= 60
                and not previous.endswith((".", "!", "?", ":", ";"))
                and not current.endswith((".", "!", "?", ":", ";"))
                and not bullet_pattern.match(current)
                and cur_words[0][:1].islower()
            ):
                merged_paragraphs[-1] = normalize_whitespace(f"{previous} {current}")
                continue
        merged_paragraphs.append(current)

    return "\n\n".join(merged_paragraphs)


def repair_hungarian_bullet_particles(text: str) -> str:
    lines = text.splitlines()
    if not lines:
        return text

    bullet_head = re.compile(r"^\s*[\u2022\u25CF\uF0B7\-*\u2023]\s+")
    particles = ("át", "be", "ki", "el", "meg", "össze", "vissza")
    verbs = (
        "Tekintse",
        "Mutassa",
        "Vizsgálja",
        "Értékelje",
        "Készítse",
        "Térjen",
        "Dolgozzon",
        "Ismertesse",
        "Valósítson",
        "Hasonlítsa",
        "Határozza",
    )

    fixed_lines = []
    previous_was_bullet_only = False
    for line in lines:
        fixed = line
        bullet_only = bool(re.match(r"^\s*[\u2022\u25CF\uF0B7\-*\u2023]\s*$", line))
        should_fix = bullet_head.match(line) or (previous_was_bullet_only and line.strip())
        if should_fix:
            for verb in verbs:
                for particle in particles:
                    fixed = re.sub(rf"\b{verb}{particle}\b", f"{verb} {particle}", fixed)
        fixed_lines.append(fixed)
        previous_was_bullet_only = bullet_only
    return "\n".join(fixed_lines)


ACCENT_MARK_MAP = {
    ("acute", "a"): "á",
    ("acute", "A"): "Á",
    ("acute", "e"): "é",
    ("acute", "E"): "É",
    ("acute", "i"): "í",
    ("acute", "I"): "Í",
    ("acute", "o"): "ó",
    ("acute", "O"): "Ó",
    ("acute", "u"): "ú",
    ("acute", "U"): "Ú",
    ("double", "o"): "ő",
    ("double", "O"): "Ő",
    ("double", "u"): "ű",
    ("double", "U"): "Ű",
    ("umlaut", "o"): "ö",
    ("umlaut", "O"): "Ö",
    ("umlaut", "u"): "ü",
    ("umlaut", "U"): "Ü",
}

ACCENT_KIND_MAP = {
    "´": "acute",
    "`": "acute",
    "\u0301": "acute",
    "˝": "double",
    "\u030B": "double",
    "¨": "umlaut",
    "\u0308": "umlaut",
}


def repair_extracted_accents(text: str) -> str:
    text = unicodedata.normalize("NFKC", text)

    def map_accent(mark: str, char: str) -> str:
        accent_kind = ACCENT_KIND_MAP.get(mark)
        if not accent_kind:
            return ""
        return ACCENT_MARK_MAP.get((accent_kind, char), "")

    def replace_mark_before(match: re.Match[str]) -> str:
        mark = match.group("mark")
        char = match.group("char")
        replaced = map_accent(mark, char)
        return replaced if replaced else f"{mark}{char}"

    def replace_mark_after(match: re.Match[str]) -> str:
        char = match.group("char")
        mark = match.group("mark")
        replaced = map_accent(mark, char)
        return replaced if replaced else f"{char}{mark}"

    def replace_split_word_accent(match: re.Match[str]) -> str:
        prev = match.group("prev")
        mark = match.group("mark")
        char = match.group("char")
        replaced = map_accent(mark, char)
        return f"{prev}{replaced}" if replaced else f"{prev}{mark}{char}"

    # PDFs often emit: letter + space + accent-mark + vowel (e.g. Czak ́o).
    text = re.sub(
        r"(?P<prev>[A-Za-z])\s+(?P<mark>[´˝¨`\u0301\u030B\u0308])\s*(?P<char>[A-Za-z])",
        replace_split_word_accent,
        text,
    )

    # Most problematic PDFs emit standalone accent marks before the vowel.
    text = re.sub(
        r"(?P<mark>[´˝¨`\u0301\u030B\u0308])\s*(?P<char>[A-Za-z])",
        replace_mark_before,
        text,
    )
    # Some generators place accent marks after the vowel instead.
    text = re.sub(
        r"(?P<char>[A-Za-z])\s*(?P<mark>[´˝¨`\u0301\u030B\u0308])",
        replace_mark_after,
        text,
    )

    accented_vowels = "ÁÉÍÓÖŐÚÜŰáéíóöőúüű"
    # Fix word splits that remain as consonant + space + accented-vowel.
    text = re.sub(
        rf"(?P<prev>[A-Za-z])\s+(?P<vowel>[{accented_vowels}])(?=[A-Za-z{accented_vowels}])",
        r"\g<prev>\g<vowel>",
        text,
    )
    # Also merge common end-of-word cases (e.g. Czak ó -> Czakó, hallgat ó -> hallgató).
    text = re.sub(
        rf"\b(?P<stem>[A-Za-z]{{3,}})\s+(?P<vowel>[{accented_vowels}])\b",
        r"\g<stem>\g<vowel>",
        text,
    )
    return unicodedata.normalize("NFC", text)


def normalize_for_matching(text: str) -> str:
    text = strip_accents(text)
    text = text.lower()
    return re.sub(r"\s+", " ", text).strip()


def clean_line(line: str) -> str:
    return normalize_whitespace(line).strip(" .:-")


def looks_like_noise(line: str) -> bool:
    return any(pattern.search(line) for pattern in NOISE_LINE_PATTERNS)


def looks_like_task_label(line: str) -> bool:
    normalized = normalize_for_matching(line)
    if any(marker in normalized for marker in STUDENT_NAME_EXCLUDE_MARKERS):
        return True

    compact = re.sub(r"[^A-Za-z\u00C0-\u017F]", "", line)
    if compact and compact.upper() == compact and len(compact) >= 10:
        return True

    return False


def default_input_dir() -> Path:
    script_path = Path(__file__).resolve()
    candidates = [Path("/data")]

    # Handle both deep repository paths and shallow container paths (e.g. /app/review_pdfs.py).
    if len(script_path.parents) >= 3:
        candidates.append(script_path.parents[2] / "data" / "td")

    candidates.append(Path.cwd())
    for candidate in candidates:
        if candidate.exists() and any(candidate.glob("*.pdf")):
            return candidate
    return Path.cwd()


def normalize_extracted_page(text: str) -> str:
    repaired = unicodedata.normalize("NFC", repair_extracted_accents(text or ""))
    cleaned = normalize_layout_spacing(repaired)
    return normalize_whitespace(repair_hungarian_bullet_particles(cleaned))


def extract_pdf_text_fitz(pdf_path: Path, verbose: int = 0) -> tuple[int, str, list[str]]:
    fitz = importlib.import_module("fitz")
    document = fitz.open(str(pdf_path))
    page_texts = []
    for page in document:
        text_mode = normalize_extracted_page(page.get_text("text", sort=True) or "")

        # Blocks mode is more robust when PDF text is flattened into very few lines.
        blocks = page.get_text("blocks", sort=True) or []
        block_chunks = []
        for block in blocks:
            if len(block) < 5:
                continue
            block_text = normalize_extracted_page(str(block[4]))
            if block_text:
                block_chunks.append(block_text)
        blocks_mode = normalize_whitespace("\n\n".join(block_chunks))

        text_lines = [line for line in text_mode.splitlines() if line.strip()]
        blocks_lines = [line for line in blocks_mode.splitlines() if line.strip()]

        text_words = len(re.findall(r"\b[\w-]+\b", text_mode, flags=re.UNICODE))
        text_long_line = max((len(line) for line in text_lines), default=0)

        # Flattened signal: very few lines despite many words or unusually long line.
        text_looks_flattened = (
            (len(text_lines) <= 3 and text_words >= 30)
            or text_long_line > 220
        )

        chosen_text = text_mode
        chosen_mode = "fitz-text"
        if blocks_mode and (text_looks_flattened or len(blocks_lines) > len(text_lines) + 2):
            chosen_text = blocks_mode
            chosen_mode = "fitz-blocks"

        if verbose >= 2:
            print(
                "[verbose:2] "
                f"{pdf_path.name}: page {page.number + 1} extraction={chosen_mode} "
                f"text_lines={len(text_lines)} blocks_lines={len(blocks_lines)} "
                f"words={text_words} longest_line={text_long_line}"
            )

        page_texts.append(chosen_text)
    document.close()
    full_text = normalize_whitespace("\n\n".join(page_texts))
    return len(page_texts), full_text, page_texts


def extract_pdf_text_pdfplumber(pdf_path: Path, verbose: int = 0) -> tuple[int, str, list[str]]:
    try:
        pdfplumber = importlib.import_module("pdfplumber")
    except ImportError as exc:
        raise SystemExit(
            "pdfplumber is required for --pdf-extractor pdfplumber. "
            "Install it with: pip install pdfplumber"
        ) from exc
    page_texts = []
    with pdfplumber.open(str(pdf_path)) as document:
        for page_index, page in enumerate(document.pages, start=1):
            text = page.extract_text(layout=False, x_tolerance=2, y_tolerance=3) or ""

            # Fallback to words reconstruction when layout text cannot be parsed.
            if len(text.strip()) < 30:
                words = page.extract_words(use_text_flow=False, keep_blank_chars=False, x_tolerance=2, y_tolerance=3) or []
                lines: dict[int, list[tuple[float, str]]] = {}
                for word in words:
                    top_bucket = int(round(float(word.get("top", 0.0)) / 2.5))
                    x0 = float(word.get("x0", 0.0))
                    token = str(word.get("text", "")).strip()
                    if token:
                        lines.setdefault(top_bucket, []).append((x0, token))
                rendered_lines = []
                for bucket in sorted(lines):
                    rendered_lines.append(" ".join(token for _, token in sorted(lines[bucket], key=lambda item: item[0])))
                text = "\n".join(rendered_lines)

            normalized = normalize_extracted_page(text)
            if verbose >= 2:
                line_count = len([line for line in normalized.splitlines() if line.strip()])
                word_count = len(re.findall(r"\b[\w-]+\b", normalized, flags=re.UNICODE))
                print(
                    "[verbose:2] "
                    f"{pdf_path.name}: page {page_index} extraction=pdfplumber "
                    f"lines={line_count} words={word_count}"
                )
            page_texts.append(normalized)
    full_text = normalize_whitespace("\n\n".join(page_texts))
    return len(page_texts), full_text, page_texts


def extract_pdf_text_pypdf(pdf_path: Path, verbose: int = 0) -> tuple[int, str, list[str]]:
    reader = PdfReader(str(pdf_path), strict=False)
    page_texts = [normalize_extracted_page(page.extract_text() or "") for page in reader.pages]
    if verbose >= 2:
        print(f"[verbose:2] {pdf_path.name}: extraction=pypdf pages={len(page_texts)}")
    full_text = normalize_whitespace("\n\n".join(page_texts))
    return len(reader.pages), full_text, page_texts


def extract_pdf_text(pdf_path: Path, verbose: int = 0, extractor: str = "auto") -> tuple[int, str, list[str]]:
    logger = logging.getLogger("pdf-extractor")
    pipelines: list[str]
    if extractor == "auto":
        pipelines = ["fitz", "pdfplumber", "pypdf"]
    else:
        pipelines = [extractor]

    last_error = None
    for pipeline in pipelines:
        try:
            if pipeline == "fitz":
                return extract_pdf_text_fitz(pdf_path, verbose)
            if pipeline == "pdfplumber":
                return extract_pdf_text_pdfplumber(pdf_path, verbose)
            if pipeline == "pypdf":
                return extract_pdf_text_pypdf(pdf_path, verbose)
            raise ValueError(f"Unsupported extractor mode: {pipeline}")
        except Exception as exc:
            last_error = exc
            if extractor == "auto":
                logger.warning(
                    "%s extraction failed for %s, trying next fallback: %s",
                    pipeline,
                    pdf_path,
                    exc,
                )
                continue
            raise

    raise RuntimeError(f"All PDF extraction backends failed for {pdf_path}: {last_error}")


def detect_language(text: str) -> str:
    normalized = normalize_for_matching(text)
    hu_score = sum(normalized.count(token) for token in LANGUAGE_PATTERNS["hu"])
    en_score = sum(normalized.count(token) for token in LANGUAGE_PATTERNS["en"])

    hu_chars = len(re.findall(r"[\u00E1\u00E9\u00ED\u00F3\u00F6\u0151\u00FA\u00FC\u0171]", text.lower()))
    hu_score += hu_chars // 5

    if hu_score == 0 and en_score == 0:
        return "unknown"
    if hu_score >= en_score + 1:
        return "hu"
    if en_score >= hu_score + 1:
        return "en"
    return "hu" if hu_chars > 0 else "en"


def detect_heading(lines: list[str]) -> str:
    for line in lines[:10]:
        normalized = normalize_for_matching(line)
        if any(marker in normalized for marker in HEADING_MARKERS):
            return clean_line(line)

    # Fuzzy fallback: if heading contains a typo, still capture it for reporting.
    heading_keywords = (
        "feladat",
        "feladatkiiras",
        "diplomaterv",
        "diplomatervezesi",
        "szakdolgozat",
        "thesis",
        "assignment",
        "description",
        "msc",
        "bsc",
    )

    best_line = ""
    best_score = 0.0
    for line in lines[:12]:
        clean = clean_line(line)
        normalized = normalize_for_matching(clean)
        if not normalized:
            continue
        if not any(keyword in normalized for keyword in heading_keywords):
            continue
        score = max((SequenceMatcher(None, normalized, marker).ratio() for marker in HEADING_MARKERS), default=0.0)
        if score > best_score:
            best_score = score
            best_line = clean

    if best_line and best_score >= 0.72:
        return best_line

    return ""


def heading_spelling_ok(heading: str) -> bool:
    normalized = normalize_for_matching(heading)
    return bool(normalized) and any(marker in normalized for marker in HEADING_MARKERS)


def infer_degree_level(heading: str) -> str:
    normalized = normalize_for_matching(heading)
    has_msc = (
        "diplomatervezesi" in normalized
        or "diplomaterv" in normalized
        or "msc" in normalized
    )
    has_bsc = "szakdolgozati" in normalized or "szakdolgozat" in normalized or "bsc" in normalized

    if has_msc and has_bsc:
        return "MSc/BSc"
    if has_msc:
        return "MSc"
    if has_bsc:
        return "BSc"
    return "unknown"


def extract_student(lines: list[str]) -> str:
    for line in lines[:20]:
        clean = clean_line(line)
        normalized = normalize_for_matching(clean)
        if looks_like_task_label(clean) or looks_like_noise(clean):
            continue
        for marker in STUDENT_MARKERS:
            if marker in normalized:
                marker_index = normalized.find(marker)
                prefix = clean[:marker_index].strip(" ,-:")
                if prefix and len(prefix.split()) >= 2:
                    return prefix

    for index in range(1, len(lines)):
        previous_line = clean_line(lines[index - 1])
        current_line = normalize_for_matching(lines[index])
        if any(marker in current_line for marker in STUDENT_MARKERS) and previous_line:
            if looks_like_task_label(previous_line) or looks_like_noise(previous_line):
                continue
            return previous_line

    for line in lines[:20]:
        clean = clean_line(line)
        if re.search(r"^[A-Z][A-Za-z\u00C0-\u017F'\- ]{4,}$", clean):
            if not looks_like_noise(clean) and not looks_like_task_label(clean) and len(clean.split()) >= 2:
                return clean
    return ""


def title_from_filename(pdf_path: Path) -> str:
    stem = pdf_path.stem
    stem = re.sub(r"-Feladatkiiras-\d+$", "", stem, flags=re.IGNORECASE)
    title = stem.replace("-", " ").strip()
    return title[:1].upper() + title[1:] if title else ""


def extract_title(lines: list[str], pdf_path: Path) -> str:
    marker_index = None
    for index, raw_line in enumerate(lines[:20]):
        norm = normalize_for_matching(raw_line)
        if any(marker in norm for marker in STUDENT_MARKERS):
            marker_index = index
            break

    windows = []
    if marker_index is not None:
        windows.append(range(marker_index + 1, min(marker_index + 7, len(lines))))
    windows.append(range(0, min(24, len(lines))))

    def is_title_candidate(line: str) -> bool:
        normalized = normalize_for_matching(line)
        if not line or looks_like_noise(line):
            return False
        if any(marker in normalized for marker in HEADING_MARKERS):
            return False
        if any(marker in normalized for marker in STUDENT_MARKERS):
            return False
        if normalized.startswith("tanszeki konzulens") or normalized.startswith("supervisor"):
            return False
        if normalized.startswith("academic supervisor") or normalized.startswith("external supervisor"):
            return False
        word_count = len(line.split())
        if len(line) > 120 or word_count < 3 or word_count > 14:
            return False
        if line.endswith((".", ":", "!", "?")):
            return False
        return True

    def collect_title_from(start_index: int, end_index: int) -> str:
        allowed_lowercase_starters = {
            "for",
            "with",
            "in",
            "of",
            "to",
            "and",
            "on",
            "using",
            "towards",
            "toward",
        }

        def is_short_title_continuation(line: str) -> bool:
            if not line or looks_like_noise(line):
                return False
            if line.endswith((".", "!", "?", ":", ";")):
                return False
            words = line.split()
            if not (1 <= len(words) <= 3):
                return False
            if not line[:1].islower():
                return False
            normalized = normalize_for_matching(line)
            if any(marker in normalized for marker in STUDENT_MARKERS):
                return False
            if any(marker in normalized for marker in HEADING_MARKERS):
                return False
            return True

        parts = []
        for idx in range(start_index, end_index):
            line = clean_line(lines[idx])
            standalone_candidate = is_title_candidate(line)
            continuation_candidate = bool(parts) and is_short_title_continuation(line)
            if not standalone_candidate and not continuation_candidate:
                if parts:
                    break
                continue

            word_count = len(line.split())
            first_token = line.split()[0].lower() if line.split() else ""
            is_lower_start = bool(line[:1]) and line[:1].islower()

            if parts:
                # Stop if this already looks like prose instead of a title continuation.
                if "," in line or ";" in line:
                    break
                if standalone_candidate and word_count > 10:
                    break
                if is_lower_start and first_token not in allowed_lowercase_starters and word_count > 3:
                    break

            parts.append(line)
            if len(parts) >= 4:
                break
        return normalize_whitespace(" ".join(parts))

    for window in windows:
        title = collect_title_from(window.start, window.stop)
        if title:
            return title

    return title_from_filename(pdf_path)


def extract_advisor(text: str) -> str:
    for line in text.splitlines():
        stripped_line = strip_accents(line)
        for pattern in ADVISOR_PATTERNS:
            match = re.search(pattern, stripped_line, re.IGNORECASE)
            if match:
                start, end = match.span(1)
                return clean_line(line[start:end])
    return ""


def extract_date(text: str) -> str:
    for pattern in DATE_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return clean_line(match.group(0))
    return ""


def merge_detected_title_occurrences(full_text: str, title: str) -> str:
    if not full_text or not title:
        return full_text

    title_words = title.split()
    if len(title_words) < 3:
        return full_text

    # Match title across arbitrary whitespace/newlines and normalize to one clean line.
    title_pattern = r"(?is)" + r"\s+".join(re.escape(word) for word in title_words)

    head_limit = min(len(full_text), 2600)
    head = full_text[:head_limit]
    tail = full_text[head_limit:]
    normalized_head, _ = re.subn(title_pattern, title, head, count=3)
    return normalized_head + tail


def count_bullets(lines: Iterable[str]) -> int:
    count = 0
    for line in lines:
        if re.match(r"^\s*[\u2022\u25CF\uF0B7\-*\u2023]\s+", line):
            count += 1
    return count


def intro_text(full_text: str, language: str) -> str:
    normalized = strip_accents(full_text)
    split_patterns = [
        r"A hallgato feladatanak a kovetkezokre kell kiterjednie:",
        r"A hallgato feladatanak a kovetkezokre kell kiternie:",
        r"The student'?s tasks include:",
        r"The tasks are as follows:",
    ]
    if language == "en":
        split_patterns = split_patterns[2:] + split_patterns[:2]
    for pattern in split_patterns:
        parts = re.split(pattern, normalized, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2:
            return normalize_whitespace(parts[0])
    return normalize_whitespace(full_text)


def keyword_present(text: str, keywords: tuple[str, ...]) -> bool:
    normalized = normalize_for_matching(text)
    return any(keyword in normalized for keyword in keywords)


def build_issue_list(checks: dict[str, bool]) -> list[str]:
    issues = []
    if not checks["heading_detected"]:
        issues.append("missing recognizable heading")
    elif not checks["heading_spelling_ok"]:
        issues.append("heading detected but likely misspelled")
    if not checks["student_name_detected"]:
        issues.append("missing student name")
    if not checks["title_detected"]:
        issues.append("missing thesis title")
    if not checks["advisor_detected"]:
        issues.append("missing advisor")
    if not checks["date_detected"]:
        issues.append("missing date")
    if not checks["page_count_reasonable"]:
        issues.append("unexpected page count")
    if not checks["task_list_present"]:
        issues.append("task list too short or missing")
    if not checks["objective_paragraph_present"]:
        issues.append("objective/problem statement too short")
    if not checks["literature_task_present"]:
        issues.append("no literature/background task detected")
    if not checks["implementation_task_present"]:
        issues.append("no implementation/modeling task detected")
    if not checks["evaluation_task_present"]:
        issues.append("no evaluation/validation task detected")
    if not checks["report_task_present"]:
        issues.append("no report/presentation deliverable detected")
    if not checks["task_count_reasonable"]:
        issues.append("task list length outside expected range")
    return issues


def compute_score(check_ids: list[str], checks: dict[str, bool]) -> int:
    passed = sum(1 for check_id in check_ids if checks[check_id])
    return round(passed * 100 / len(check_ids)) if check_ids else 0


class SpellcheckEngine:
    def __init__(self, mode: str, language_arg: str):
        self.mode = mode
        self.language_arg = language_arg
        self._cache: dict[str, object] = {}
        self._cache_errors: dict[str, str] = {}
        self._language_tool_module = None

        if self.mode != "none":
            try:
                self._language_tool_module = importlib.import_module("language_tool_python")
            except ImportError as exc:
                raise SystemExit(
                    "language_tool_python is required for spell/grammar check. "
                    "Install it with: pip install language_tool_python"
                ) from exc

    def _resolve_language(self, detected_language: str) -> str:
        if self.language_arg.lower() != "auto":
            return self.language_arg
        if detected_language == "en":
            return "en-US"
        return "hu-HU"

    def _get_tool(self, language_code: str):
        if language_code in self._cache:
            return self._cache[language_code]

        if self._language_tool_module is None:
            self._cache_errors[language_code] = "LanguageTool module is not loaded."
            return None

        try:
            if self.mode == "auto":
                tool = self._language_tool_module.LanguageToolPublicAPI(language_code)
            elif self.mode == "languagetool":
                tool = self._language_tool_module.LanguageTool(language_code)
            else:
                tool = None
        except Exception as exc:
            self._cache_errors[language_code] = str(exc)
            return None

        self._cache[language_code] = tool
        return tool

    def check(self, text: str, detected_language: str) -> SpellcheckSummary:
        if self.mode == "none":
            return SpellcheckSummary("disabled", 0, 0, 0, "Spell and grammar check disabled.")

        language_code = self._resolve_language(detected_language)
        tool = self._get_tool(language_code)
        if tool is None:
            error = self._cache_errors.get(language_code, "Spellcheck engine was not initialized.")
            return SpellcheckSummary("failed", 0, 0, 0, f"Spell/grammar check unavailable for {language_code}: {error}")

        chunks = []
        current = []
        current_len = 0
        for paragraph in text.split("\n\n"):
            paragraph = paragraph.strip()
            if not paragraph:
                continue
            if current_len + len(paragraph) > 3500 and current:
                chunks.append("\n\n".join(current))
                current = [paragraph]
                current_len = len(paragraph)
            else:
                current.append(paragraph)
                current_len += len(paragraph)
        if current:
            chunks.append("\n\n".join(current))

        try:
            matches = []
            for chunk in chunks[:6]:
                matches.extend(tool.check(chunk))
        except Exception as exc:
            return SpellcheckSummary("failed", 0, 0, 0, f"Spell/grammar check failed: {exc}")

        filtered = []
        for match in matches:
            category = (getattr(getattr(match, "category", None), "id", "") or "").upper()
            if category in {"TYPOGRAPHY", "WHITESPACE"}:
                continue
            filtered.append(match)

        grammar_count = 0
        spelling_count = 0
        for match in filtered:
            issue_type = (getattr(match, "ruleIssueType", "") or "").lower()
            if issue_type == "misspelling":
                spelling_count += 1
            else:
                grammar_count += 1

        examples = []
        for match in filtered[:5]:
            context = re.sub(r"\s+", " ", getattr(match, "context", "")).strip()
            message = getattr(match, "message", "issue")
            examples.append(f"{message}: {context[:110]}")

        summary = "; ".join(examples) if examples else "No significant language issues returned by LanguageTool."
        return SpellcheckSummary(
            status="done",
            issue_count=len(filtered),
            grammar_issue_count=grammar_count,
            spelling_issue_count=spelling_count,
            summary=summary,
        )


def call_openai_text_review(
    *,
    mode: str,
    base_url: str,
    api_key: str,
    model: str,
    language: str,
    text: str,
    title: str,
    system_prompt: str,
) -> LlmCheckSummary:
    if mode == "none":
        return LlmCheckSummary("disabled", 0, "", "LLM review disabled.")

    if not system_prompt.strip():
        return LlmCheckSummary("skipped", 0, "", "LLM review skipped: --llm-prompt-file was not specified.")

    if not base_url:
        return LlmCheckSummary("failed", 0, "", "Missing --openai-base-url for LLM review.")
    if not api_key:
        return LlmCheckSummary("failed", 0, "", "Missing OpenAI API key for LLM review.")

    endpoint = resolve_openai_endpoint(base_url)
    user_prompt = build_llm_user_text(language, title, text)

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
            {"role": "assistant", "content": "<think></think>"},
        ],
        "temperature": 0.1,
    }

    request = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        return LlmCheckSummary("failed", 0, "", f"LLM review HTTP error {exc.code}: {details[:180]}")
    except Exception as exc:
        return LlmCheckSummary("failed", 0, "", f"LLM review failed: {exc}")

    try:
        parsed = json.loads(raw)
        content = parsed["choices"][0]["message"]["content"]
        content = re.sub(r"^\s*<think>\s*</think>\s*", "", content, flags=re.IGNORECASE)
    except Exception as exc:
        return LlmCheckSummary("failed", 0, "", f"Invalid LLM response envelope: {exc}", raw_answer=raw[:2000])

    def extract_json_content(text: str) -> tuple[dict | None, str]:
        stripped = text.strip()
        try:
            return json.loads(stripped), stripped
        except Exception:
            pass

        decoder = json.JSONDecoder()
        for index, char in enumerate(stripped):
            if char != "{":
                continue
            try:
                obj, end = decoder.raw_decode(stripped[index:])
                return obj, stripped[index:index + end]
            except Exception:
                continue
        return None, ""

    review_json, json_text = extract_json_content(content)
    if review_json is None:
        return LlmCheckSummary(
            "failed",
            0,
            "",
            "Could not parse LLM JSON result from answer content.",
            raw_answer=content,
        )

    try:
        score = int(review_json.get("score_0_100", 0))
        verdict = str(review_json.get("verdict", "REVIEW")).upper()
        summary = str(review_json.get("summary", ""))
        findings = review_json.get("findings", [])
        if isinstance(findings, list) and findings:
            summary = f"{summary} Findings: " + " | ".join(str(item) for item in findings[:4])
        return LlmCheckSummary(
            "done",
            max(0, min(100, score)),
            verdict,
            summary[:400],
            parsed_json=json_text,
            raw_answer=content,
        )
    except Exception as exc:
        return LlmCheckSummary(
            "failed",
            0,
            "",
            f"Could not parse LLM JSON fields: {exc}",
            parsed_json=json_text,
            raw_answer=content,
        )


def is_timeout_like_llm_failure(summary: LlmCheckSummary) -> bool:
    if summary.status != "failed":
        return False
    message = (summary.summary or "").lower()
    timeout_markers = (
        "timed out",
        "timeout",
        "time out",
        "read operation timed out",
        "gateway timeout",
        "http error 504",
        "http error 524",
    )
    return any(marker in message for marker in timeout_markers)


def review_pdf(pdf_path: Path, spell_engine: SpellcheckEngine, args: argparse.Namespace) -> dict[str, object]:
    page_count, full_text, page_texts = extract_pdf_text(pdf_path, args.verbose, args.pdf_extractor)
    first_page_lines = [line for line in page_texts[0].splitlines() if clean_line(line)] if page_texts else []

    language = detect_language(full_text)
    heading = detect_heading(first_page_lines)
    student_name = extract_student(first_page_lines)
    title = extract_title(first_page_lines, pdf_path)
    full_text = merge_detected_title_occurrences(full_text, title)
    advisor = extract_advisor(full_text)
    signed_date = extract_date(full_text)
    bullets = count_bullets(page_texts[0].splitlines() if page_texts else [])
    intro = intro_text(full_text, language)
    text_word_count = len(re.findall(r"\b[\w-]+\b", full_text, flags=re.UNICODE))
    intro_char_count = len(intro)
    txt_export_text = build_txt_export_text(language, title, full_text)

    txt_output_path = write_extracted_text_file(pdf_path, txt_export_text, args.txt_output_dir)
    if txt_output_path is not None:
        vprint(args, 2, f"{pdf_path.name}: final analysis text saved to {txt_output_path}")

    keyword_source = KEYWORD_GROUPS.get(language, KEYWORD_GROUPS["en"])

    checks = {
        "heading_detected": bool(heading),
        "heading_spelling_ok": (not heading) or heading_spelling_ok(heading),
        "student_name_detected": bool(student_name),
        "title_detected": bool(title),
        "advisor_detected": bool(advisor),
        "date_detected": bool(signed_date),
        "page_count_reasonable": 1 <= page_count <= 2,
        "task_list_present": bullets >= 4,
        "objective_paragraph_present": intro_char_count >= 300,
        "literature_task_present": keyword_present(full_text, keyword_source["literature"]),
        "implementation_task_present": keyword_present(full_text, keyword_source["implementation"]),
        "evaluation_task_present": keyword_present(full_text, keyword_source["evaluation"]),
        "report_task_present": keyword_present(full_text, keyword_source["report"]),
        "task_count_reasonable": 4 <= bullets <= 8,
    }

    # If LLM review is enabled, it is responsible for language quality checks.
    if args.effective_llm_check == "openai":
        spellcheck = SpellcheckSummary(
            status="skipped",
            issue_count=0,
            grammar_issue_count=0,
            spelling_issue_count=0,
            summary="Spell/grammar check skipped because LLM review is enabled.",
        )
    else:
        spellcheck = spell_engine.check(full_text, language)
    api_key = args.openai_api_key or os.environ.get(args.openai_api_key_env, "")
    if args.effective_llm_check == "openai" and not api_key:
        api_key = "dummy"
    llm_endpoint = resolve_openai_endpoint(args.openai_base_url)

    if args.effective_llm_check == "openai":
        vprint(args, 1, f"{pdf_path.name}: starting LLM review ({args.openai_model})")
        vprint(
            args,
            2,
            (
                f"{pdf_path.name}: endpoint={llm_endpoint or '<missing>'} "
                f"language={language} text_chars={len(full_text)} "
                f"api_key_present={'yes' if bool(api_key) else 'no'}"
            ),
        )

    llm = call_openai_text_review(
        mode=args.effective_llm_check,
        base_url=args.openai_base_url,
        api_key=api_key,
        model=args.openai_model,
        language=language,
        text=full_text,
        title=title,
        system_prompt=args.llm_system_prompt,
    )

    if args.effective_llm_check == "openai" and is_timeout_like_llm_failure(llm):
        vprint(args, 1, f"{pdf_path.name}: LLM timeout-like failure, retrying once")
        retry_llm = call_openai_text_review(
            mode=args.effective_llm_check,
            base_url=args.openai_base_url,
            api_key=api_key,
            model=args.openai_model,
            language=language,
            text=full_text,
            title=title,
            system_prompt=args.llm_system_prompt,
        )
        if retry_llm.status == "done":
            vprint(args, 1, f"{pdf_path.name}: LLM retry succeeded")
            llm = retry_llm
        else:
            llm = retry_llm

    if args.effective_llm_check == "openai":
        if llm.status == "done":
            vprint(args, 1, f"{pdf_path.name}: LLM review finished status={llm.status} verdict={llm.verdict} score={llm.score}")
            vprint(
                args,
                2,
                (
                    f"{pdf_path.name}: LLM result debug: "
                    f"status={llm.status} score={llm.score} verdict={llm.verdict} "
                    f"summary={llm.summary[:260]}"
                ),
            )
            vprint(args, 2, f"{pdf_path.name}: LLM raw answer:\n{llm.raw_answer}")
            vprint(args, 2, f"{pdf_path.name}: LLM parsed JSON used:\n{llm.parsed_json}")
        else:
            vprint(args, 1, f"{pdf_path.name}: LLM review finished status={llm.status}")
            vprint(
                args,
                2,
                (
                    f"{pdf_path.name}: LLM result debug: "
                    f"status={llm.status} score={llm.score} verdict={llm.verdict or '<none>'} "
                    f"summary={llm.summary[:260]}"
                ),
            )
            if llm.raw_answer:
                vprint(args, 2, f"{pdf_path.name}: LLM raw answer:\n{llm.raw_answer}")
            if llm.parsed_json:
                vprint(args, 2, f"{pdf_path.name}: LLM parsed JSON used:\n{llm.parsed_json}")

    formal_score = compute_score([item["id"] for item in FORMAL_RUBRICS], checks)
    content_score = compute_score([item["id"] for item in CONTENT_RUBRICS], checks)
    issues = build_issue_list(checks)

    if spellcheck.issue_count >= 8:
        issues.append("many possible spelling/grammar issues")
    if spellcheck.status == "failed":
        issues.append("spell/grammar check failed")
    if llm.status == "failed":
        issues.append("LLM review failed")

    overall_status = "OK"
    critical_fields = [
        checks["student_name_detected"],
        checks["title_detected"],
        checks["advisor_detected"],
        checks["task_list_present"],
    ]
    if not all(critical_fields) or formal_score < 70 or content_score < 65 or spellcheck.issue_count >= 8:
        overall_status = "REVIEW"
    if llm.status == "done" and (llm.verdict == "REVIEW" or llm.score < 65):
        overall_status = "REVIEW"

    language_label = "hu" if language == "hu" else "en" if language == "en" else "unknown"

    return {
        "file_name": pdf_path.name,
        "file_path": str(pdf_path.resolve()),
        "detected_language": language_label,
        "degree_level": infer_degree_level(heading),
        "document_heading": heading,
        "student_name": student_name,
        "thesis_title": title,
        "advisor": advisor,
        "signed_date": signed_date,
        "page_count": page_count,
        "word_count": text_word_count,
        "intro_char_count": intro_char_count,
        "task_bullet_count": bullets,
        "formal_score": formal_score,
        "content_score": content_score,
        "spellcheck_status": spellcheck.status,
        "language_issue_count": spellcheck.issue_count,
        "grammar_issue_count": spellcheck.grammar_issue_count,
        "spelling_issue_count": spellcheck.spelling_issue_count,
        "spellcheck_summary": spellcheck.summary,
        "llm_check_status": llm.status,
        "llm_score": llm.score,
        "llm_verdict": llm.verdict,
        "llm_summary": llm.summary,
        "heading_detected": "yes" if checks["heading_detected"] else "no",
        "heading_spelling_ok": "yes" if checks["heading_spelling_ok"] else "no",
        "student_name_detected": "yes" if checks["student_name_detected"] else "no",
        "title_detected": "yes" if checks["title_detected"] else "no",
        "advisor_detected": "yes" if checks["advisor_detected"] else "no",
        "date_detected": "yes" if checks["date_detected"] else "no",
        "page_count_reasonable": "yes" if checks["page_count_reasonable"] else "no",
        "task_list_present": "yes" if checks["task_list_present"] else "no",
        "objective_paragraph_present": "yes" if checks["objective_paragraph_present"] else "no",
        "literature_task_present": "yes" if checks["literature_task_present"] else "no",
        "implementation_task_present": "yes" if checks["implementation_task_present"] else "no",
        "evaluation_task_present": "yes" if checks["evaluation_task_present"] else "no",
        "report_task_present": "yes" if checks["report_task_present"] else "no",
        "task_count_reasonable": "yes" if checks["task_count_reasonable"] else "no",
        "auto_flags": "; ".join(issues),
        "overall_status": overall_status,
        "manual_formal_review": "",
        "manual_spell_review": "",
        "manual_content_review": "",
        "manual_notes": "",
    }


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 48)


def write_workbook(rows: list[dict[str, object]], output_path: Path) -> None:
    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "reviews"

    ordered_columns = [
        "file_name",
        "detected_language",
        "degree_level",
        "document_heading",
        "student_name",
        "thesis_title",
        "advisor",
        "signed_date",
        "page_count",
        "word_count",
        "task_bullet_count",
        "formal_score",
        "content_score",
        "spellcheck_status",
        "language_issue_count",
        "grammar_issue_count",
        "spelling_issue_count",
        "llm_check_status",
        "llm_score",
        "llm_verdict",
        "overall_status",
        "auto_flags",
        "spellcheck_summary",
        "llm_summary",
        "file_path",
        "heading_detected",
        "heading_spelling_ok",
        "student_name_detected",
        "title_detected",
        "advisor_detected",
        "date_detected",
        "page_count_reasonable",
        "task_list_present",
        "objective_paragraph_present",
        "literature_task_present",
        "implementation_task_present",
        "evaluation_task_present",
        "report_task_present",
        "task_count_reasonable",
        "manual_formal_review",
        "manual_spell_review",
        "manual_content_review",
        "manual_notes",
    ]

    review_sheet.append(ordered_columns)
    for row in rows:
        review_sheet.append([row.get(column, "") for column in ordered_columns])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    warning_fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    ok_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

    for cell in review_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    status_col_index = ordered_columns.index("overall_status")
    for row in review_sheet.iter_rows(min_row=2):
        status_cell = row[status_col_index]
        status_cell.fill = ok_fill if status_cell.value == "OK" else warning_fill

    review_sheet.freeze_panes = "A2"
    review_sheet.auto_filter.ref = review_sheet.dimensions
    autosize_columns(review_sheet)

    rubric_sheet = workbook.create_sheet("rubric")
    rubric_sheet.append(["group", "check_id", "description", "automatic_pass_rule"])
    for cell in rubric_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for rubric in FORMAL_RUBRICS + CONTENT_RUBRICS:
        rubric_sheet.append([rubric["group"], rubric["id"], rubric["description"], rubric["pass_rule"]])

    autosize_columns(rubric_sheet)

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["metric", "value"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    formal_scores = [int(row["formal_score"]) for row in rows] if rows else []
    content_scores = [int(row["content_score"]) for row in rows] if rows else []
    language_issues = [int(row["language_issue_count"]) for row in rows] if rows else []
    review_count = sum(1 for row in rows if row["overall_status"] == "REVIEW")

    summary_rows = [
        ("files_reviewed", len(rows)),
        ("requires_review", review_count),
        ("average_formal_score", round(statistics.mean(formal_scores), 1) if formal_scores else 0),
        ("average_content_score", round(statistics.mean(content_scores), 1) if content_scores else 0),
        ("average_language_issue_count", round(statistics.mean(language_issues), 1) if language_issues else 0),
    ]
    for metric, value in summary_rows:
        summary_sheet.append([metric, value])

    autosize_columns(summary_sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csv(rows: list[dict[str, object]], output_path: Path) -> None:
    ordered_columns = list(rows[0].keys()) if rows else []
    if not ordered_columns:
        return
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=ordered_columns)
        writer.writeheader()
        writer.writerows(rows)


def write_extracted_text_file(pdf_path: Path, text: str, txt_output_dir: str) -> Path | None:
    if not txt_output_dir:
        return None

    target_dir = Path(txt_output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"{pdf_path.stem}.txt"
    target_path.write_text(text, encoding="utf-8")
    return target_path


def build_llm_user_text(language: str, title: str, text: str) -> str:
    text_excerpt = normalize_whitespace(text)[:12000]
    language_name = "Hungarian" if language == "hu" else "English"
    return (
        f"Task description language: {language_name}.\n"
        f"Title: {title or 'unknown'}\n"
        f"Task description text:\n{text_excerpt}"
    )


def build_txt_export_text(language: str, title: str, text: str) -> str:
    text_excerpt = normalize_whitespace(text)[:12000]
    language_name = "Hungarian" if language == "hu" else "English"
    return (
        f"Task description language: {language_name}.\n"
        f"Title: {title or 'unknown'}\n\n"
        f"Task description text:\n{text_excerpt}"
    )


def vprint(args: argparse.Namespace, level: int, message: str) -> None:
    if args.verbose >= level:
        print(f"[verbose:{level}] {message}")


def resolve_openai_endpoint(base_url: str) -> str:
    endpoint = base_url.rstrip("/")
    if not endpoint:
        return endpoint
    if endpoint.endswith("/chat/completions"):
        return endpoint
    if endpoint.endswith("/v1"):
        return f"{endpoint}/chat/completions"
    return f"{endpoint}/v1/chat/completions"

def infer_llm_mode(args: argparse.Namespace) -> str:
    # Explicit override wins.
    if args.llm_check in {"none", "openai"}:
        return args.llm_check

    # Auto mode: enable OpenAI only when the required flags are present.
    if args.openai_base_url and args.llm_prompt_file:
        return "openai"
    return "none"

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Review thesis-task PDFs and export a rubric workbook.")
    parser.add_argument("--input-dir", type=Path, default=default_input_dir(), help="Directory containing PDF files.")
    parser.add_argument("--pattern", default="*.pdf", help="Glob pattern for selecting PDF files.")
    parser.add_argument("--output", type=Path, default=Path("pdf_review.xlsx"), help="Output XLSX path.")
    parser.add_argument(
        "--txt-output-dir",
        default="",
        help="Optional directory to save extracted PDF text as .txt files using the same base filename.",
    )
    parser.add_argument(
        "--pdf-extractor",
        choices=("auto", "fitz", "pdfplumber", "pypdf"),
        default="auto",
        help="PDF text extractor backend. auto=fitz->pdfplumber->pypdf fallback chain.",
    )
    parser.add_argument(
        "--spellcheck",
        choices=("none", "auto", "languagetool"),
        default="auto",
        help="Spell+grammar mode. auto=LanguageTool public API, languagetool=local backend, none=disabled.",
    )
    parser.add_argument(
        "--language",
        default="auto",
        help="Language for spell+grammar check. Use auto, hu-HU, en-US, etc.",
    )
    parser.add_argument(
        "--llm-check",
        choices=("auto", "none", "openai"),
        default="auto",
        help="LLM review mode. Default is auto (infers OpenAI when required flags are present).",
    )
    parser.add_argument("--openai-base-url", default="", help="OpenAI-compatible base URL, e.g. https://api.openai.com")
    parser.add_argument("--openai-model", default="gpt-4o-mini", help="Model name for LLM review.")
    parser.add_argument("--openai-api-key", default="", help="API key (password/token) for LLM review.")
    parser.add_argument("--openai-api-key-env", default="OPENAI_API_KEY", help="Environment variable name for API key fallback.")
    parser.add_argument(
        "--llm-prompt-file",
        default="",
        help="Path to text file containing the SYSTEM prompt for LLM review.",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="count",
        default=0,
        help="Increase progress output. Use -v for file-level info and -vv for detailed check info.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.llm_prompt_file:
        prompt_path = Path(args.llm_prompt_file)
        try:
            args.llm_system_prompt = prompt_path.read_text(encoding="utf-8")
        except Exception as exc:
            raise SystemExit(f"Cannot read --llm-prompt-file {prompt_path}: {exc}") from exc
    else:
        args.llm_system_prompt = ""

    args.effective_llm_check = infer_llm_mode(args)

    if args.effective_llm_check == "openai" and args.spellcheck != "none":
        vprint(
            args,
            1,
            "LLM review enabled; skipping standalone spell/grammar engine (forcing --spellcheck=none).",
        )
        args.spellcheck = "none"

    vprint(args, 1, f"Input directory: {args.input_dir}")
    vprint(args, 1, f"Pattern: {args.pattern}")
    vprint(args, 1, f"Output: {args.output}")
    vprint(args, 1, f"TXT output dir: {args.txt_output_dir or '<disabled>'}")
    vprint(args, 1, f"PDF extractor: {args.pdf_extractor}")
    vprint(args, 1, f"Spellcheck mode: {args.spellcheck}, language mode: {args.language}")
    vprint(args, 1, f"LLM check mode: {args.llm_check} (effective: {args.effective_llm_check})")
    if args.effective_llm_check == "openai":
        resolved_endpoint = resolve_openai_endpoint(args.openai_base_url)
        api_key = args.openai_api_key or os.environ.get(args.openai_api_key_env, "")
        if not api_key:
            api_key = "dummy"
        vprint(args, 1, f"LLM endpoint: {resolved_endpoint or '<missing>'}")
        vprint(args, 1, f"LLM model: {args.openai_model}")
        vprint(args, 1, f"LLM api key present: {'yes' if bool(api_key) else 'no'}")
        if api_key == "dummy":
            vprint(args, 1, "LLM api key source: dummy fallback")
        vprint(args, 1, f"LLM prompt file: {args.llm_prompt_file or '<none>'}")
    pdf_files = sorted(path for path in args.input_dir.glob(args.pattern) if path.is_file())
    if not pdf_files:
        raise SystemExit(f"No PDF files found in {args.input_dir} matching {args.pattern}")
    vprint(args, 1, f"Found {len(pdf_files)} PDF files")

    spell_engine = SpellcheckEngine(args.spellcheck, args.language)
    rows = []
    for index, pdf_path in enumerate(pdf_files, start=1):
        vprint(args, 1, f"[{index}/{len(pdf_files)}] Reviewing {pdf_path.name}")
        row = review_pdf(pdf_path, spell_engine, args)
        rows.append(row)
        vprint(
            args,
            2,
            (
                f"{pdf_path.name}: lang={row.get('detected_language')} "
                f"formal={row.get('formal_score')} content={row.get('content_score')} "
                f"spell_status={row.get('spellcheck_status')} issues={row.get('language_issue_count')} "
                f"llm_status={row.get('llm_check_status')} overall={row.get('overall_status')}"
            ),
        )

    csv_output = args.output.with_suffix(".csv")
    try:
        vprint(args, 1, f"Writing workbook to {args.output}")
        write_workbook(rows, args.output)
        vprint(args, 1, f"Writing CSV to {csv_output}")
        write_csv(rows, csv_output)
    except PermissionError as exc:
        raise SystemExit(f"Cannot write output file: {exc}. Choose a writable --output path.") from exc

    print(f"Reviewed {len(rows)} files")
    print(f"Workbook: {args.output}")
    print(f"CSV: {csv_output}")


if __name__ == "__main__":
    main()
